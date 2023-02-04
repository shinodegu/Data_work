import openpyxl
import csv



with open("data.csv", "r", encoding="utf8") as file_exel:
    new_exel = csv.reader(file_exel)
    data = []
    for item in new_exel:
        data.append(item)

print(data)

for item in data:
    index_person = 1
    del item[index_person]

print(data)

wb = openpyxl.Workbook()
wb.create_sheet(title='Первый лист', index=0)

sheet = wb['Первый лист']
print(sheet)

for row_index, row in enumerate(data):
    for col_index,value in enumerate(row):
        cell = sheet.cell(row = row_index+1, column = col_index+1)
        cell.value = value

wb.save('EXEL.xlsx')